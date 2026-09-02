import csv
import os
import smtplib
from datetime import datetime, timezone
from email.message import EmailMessage
from html import escape
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.core.config import HOSPITAL_LOGO_PATH, HOSPITAL_NAME, SMTP_FROM_EMAIL, SMTP_HOST, SMTP_PASSWORD, SMTP_PORT, SMTP_USE_TLS, SMTP_USERNAME
from app.models.prediction import Prediction
from app.schemas.prediction import EmailReportRequest, PredictionSearchParams
from app.services.activity_service import record_activity
from app.services.explainability_service import build_explanation
from app.services.prediction_service import search_predictions


HEADER_FILL = PatternFill("solid", fgColor="123047")
SUBHEADER_FILL = PatternFill("solid", fgColor="DDF2F4")
HIGH_FILL = PatternFill("solid", fgColor="FFD9DE")
MEDIUM_FILL = PatternFill("solid", fgColor="FFF0C7")
LOW_FILL = PatternFill("solid", fgColor="DDF6E8")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9E6EC"))


def get_prediction_or_none(db: Session, prediction_id: int) -> Prediction | None:
    return db.query(Prediction).filter(Prediction.id == prediction_id).first()


def _percentage(value: float | None) -> str:
    return f"{value * 100:.1f}%" if value is not None else "Not available"


def _report_rows(prediction: Prediction) -> list[tuple[str, str]]:
    return [
        ("Patient name", prediction.patient_name or "Not provided"),
        ("Patient ID", prediction.patient_id or "Not provided"),
        ("Diagnosis", prediction.diagnosis or "Not recorded"),
        ("Report status", prediction.status or "draft"),
        ("Follow-up date", prediction.follow_up_date.isoformat() if prediction.follow_up_date else "Not scheduled"),
        ("Age", str(prediction.age or "Not provided")),
        ("Gender", "Male" if prediction.gender == 1 else "Female" if prediction.gender == 0 else "Not provided"),
        ("Risk level", prediction.risk or "Unknown"),
        ("Final probability", _percentage(prediction.final_probability)),
        ("Clinical probability", _percentage(prediction.clinical_probability)),
        ("Keystroke probability", _percentage(prediction.keystroke_probability)),
        ("Doctor notes", prediction.doctor_notes or "No notes recorded"),
        ("Recommendation", prediction.recommendation or "No recommendation recorded"),
    ]


def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=20, textColor=colors.HexColor("#123047"), spaceAfter=4))
    styles.add(ParagraphStyle(name="ReportSection", parent=styles["Heading3"], fontName="Helvetica-Bold", fontSize=11, textColor=colors.HexColor("#006D7C"), spaceBefore=14, spaceAfter=6))
    styles.add(ParagraphStyle(name="ReportSmall", parent=styles["BodyText"], fontSize=8, leading=11, textColor=colors.HexColor("#506A78")))
    return styles


def _safe_paragraph(value: object, style) -> Paragraph:
    return Paragraph(escape(str(value)).replace("\n", "<br/>") or "—", style)


def build_pdf(prediction: Prediction) -> bytes:
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.55 * inch, leftMargin=0.55 * inch, topMargin=0.5 * inch, bottomMargin=0.55 * inch)
    styles = _pdf_styles()
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    explanation = build_explanation(prediction)
    logo_flowable = Image(HOSPITAL_LOGO_PATH, width=1.25 * inch, height=0.48 * inch) if HOSPITAL_LOGO_PATH and os.path.exists(HOSPITAL_LOGO_PATH) else Paragraph(f"<b>{escape(HOSPITAL_NAME)}</b>", styles["ReportTitle"])
    story = [
        logo_flowable,
        Paragraph("Clinical stroke-risk assessment report", styles["Heading2"]),
        Paragraph(f"Generated {generated_at} · Report #{prediction.id} · Hospital clinical intelligence workspace", styles["ReportSmall"]),
        Spacer(1, 0.12 * inch),
    ]

    summary_rows = [[_safe_paragraph("Patient and report details", styles["BodyText"]), ""]]
    summary_rows.extend([[_safe_paragraph(label, styles["BodyText"]), _safe_paragraph(value, styles["BodyText"])] for label, value in _report_rows(prediction)])
    summary_rows.extend([
        [_safe_paragraph("Average glucose", styles["BodyText"]), _safe_paragraph(prediction.avg_glucose_level or "Not provided", styles["BodyText"])],
        [_safe_paragraph("BMI", styles["BodyText"]), _safe_paragraph(prediction.bmi or "Not provided", styles["BodyText"])],
        [_safe_paragraph("Hypertension", styles["BodyText"]), _safe_paragraph("Yes" if prediction.hypertension == 1 else "No", styles["BodyText"])],
        [_safe_paragraph("Heart disease", styles["BodyText"]), _safe_paragraph("Yes" if prediction.heart_disease == 1 else "No", styles["BodyText"])],
        [_safe_paragraph("Smoking status", styles["BodyText"]), _safe_paragraph("Current smoker" if prediction.smoking_status == 1 else "Not currently smoking", styles["BodyText"])],
    ])
    summary_table = PdfTable(summary_rows, colWidths=[1.85 * inch, 4.95 * inch], repeatRows=1)
    summary_table.setStyle(TableStyle([
        ("SPAN", (0, 0), (-1, 0)),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DDF2F4")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D8E6EC")),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F0F7F8")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)

    story.append(Paragraph("Prediction summary", styles["ReportSection"]))
    prediction_table = PdfTable([
        ["Clinical probability", "Keystroke probability", "Final probability", "Risk badge"],
        [_percentage(prediction.clinical_probability), _percentage(prediction.keystroke_probability), _percentage(prediction.final_probability), prediction.risk or "Unknown"],
    ], colWidths=[1.7 * inch] * 4)
    prediction_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123047")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F7FBFC")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D8E6EC")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(prediction_table)

    story.append(Paragraph("Explainable AI", styles["ReportSection"]))
    story.append(_safe_paragraph(explanation["clinical_explanation"], styles["BodyText"]))
    factor_rows = [["Feature", "Contribution", "Direction", "Observed value"]]
    factor_rows.extend([[factor["feature"], f"{factor['contribution_percentage']:.1f}%", factor["direction"], str(factor["value"] or "—")] for factor in explanation["feature_importance"][:8]])
    factor_table = PdfTable(factor_rows, colWidths=[2.05 * inch, 1.05 * inch, 1.15 * inch, 2.55 * inch], repeatRows=1)
    factor_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123047")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D8E6EC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(factor_table)
    story.append(Spacer(1, 0.08 * inch))
    story.append(_safe_paragraph(f"Method: {explanation['method']}. Contributions are model-local associations, not causal effects.", styles["ReportSmall"]))

    story.append(Paragraph("Doctor notes and recommendation", styles["ReportSection"]))
    notes_table = PdfTable([
        ["Diagnosis", prediction.diagnosis or "Not recorded"],
        ["Notes", prediction.doctor_notes or "No notes recorded"],
        ["Recommendation", prediction.recommendation or "No recommendation recorded"],
        ["Follow-up date", prediction.follow_up_date.isoformat() if prediction.follow_up_date else "Not scheduled"],
    ], colWidths=[1.35 * inch, 5.45 * inch])
    notes_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D8E6EC")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F0F7F8")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(notes_table)
    story.append(Spacer(1, 0.16 * inch))
    story.append(Paragraph("QR reference", styles["ReportSection"]))
    qr_code = qr.QrCodeWidget(f"PreStrokeNet report reference: /predictions/{prediction.id}")
    bounds = qr_code.getBounds()
    qr_drawing = Drawing(72, 72, transform=[72 / max(bounds[2] - bounds[0], 1), 0, 0, 72 / max(bounds[3] - bounds[1], 1), 0, 0])
    qr_drawing.add(qr_code)
    story.append(qr_drawing)
    story.append(_safe_paragraph(f"PreStrokeNet report reference: /predictions/{prediction.id}", styles["ReportSmall"]))
    story.append(Spacer(1, 0.18 * inch))
    story.append(_safe_paragraph("This report supports, but does not replace, professional clinical judgment and local protocols. Verify all inputs and recommendations before acting.", styles["ReportSmall"]))
    document.build(story)
    return buffer.getvalue()


def _prediction_export_headers() -> list[str]:
    return ["id", "patient_name", "patient_id", "diagnosis", "status", "follow_up_date", "age", "gender", "created_at", "final_probability", "clinical_probability", "keystroke_probability", "risk", "doctor_notes", "recommendation"]


def _prediction_export_row(item: Prediction) -> list[object]:
    return [
        item.id,
        item.patient_name or "",
        item.patient_id or "",
        item.diagnosis or "",
        item.status or "draft",
        item.follow_up_date.isoformat() if item.follow_up_date else "",
        item.age if item.age is not None else "",
        "Male" if item.gender == 1 else "Female" if item.gender == 0 else "",
        item.created_at.isoformat() if item.created_at else "",
        item.final_probability,
        item.clinical_probability,
        item.keystroke_probability,
        item.risk or "",
        item.doctor_notes or "",
        item.recommendation or "",
    ]


def _style_prediction_sheet(sheet, headers: list[str], rows: list[list[object]]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 2)}"
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_index, row in enumerate(rows, start=2):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.border = THIN_BORDER
            if headers[column_index - 1].endswith("probability"):
                cell.number_format = "0.0%"
            if headers[column_index - 1] == "risk":
                risk = str(value).lower()
                cell.fill = HIGH_FILL if risk == "high" else MEDIUM_FILL if risk == "medium" else LOW_FILL if risk == "low" else PatternFill(fill_type=None)
    if rows:
        ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
        table = Table(displayName=f"PredictionsTable{sheet.title.replace(' ', '')}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)
    for column_cells in sheet.columns:
        width = min(max(max(len(str(cell.value or "")) for cell in column_cells) + 2, 12), 38)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def build_excel(prediction: Prediction) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["PreStrokeNet clinical stroke-risk assessment"])
    summary["A1"].fill = HEADER_FILL
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    summary.merge_cells("A1:B1")
    for label, value in _report_rows(prediction):
        summary.append([label, value])
    for cell in summary["A"]:
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True, color="123047")
        cell.border = THIN_BORDER
    for cell in summary["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 80

    predictions = workbook.create_sheet("Predictions")
    headers = _prediction_export_headers()
    predictions.append(headers)
    predictions.append(_prediction_export_row(prediction))
    _style_prediction_sheet(predictions, headers, [_prediction_export_row(prediction)])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_bulk_excel(db: Session, params: PredictionSearchParams) -> bytes:
    query = db.query(Prediction)
    if params.q:
        search_term = f"%{params.q.lower()}%"
        query = query.filter(or_(func.lower(Prediction.patient_name).like(search_term), func.lower(Prediction.patient_id).like(search_term)))
    if params.risk:
        query = query.filter(func.lower(Prediction.risk) == params.risk.lower())
    if params.min_age is not None:
        query = query.filter(Prediction.age >= params.min_age)
    if params.max_age is not None:
        query = query.filter(Prediction.age <= params.max_age)
    if params.gender is not None:
        query = query.filter(Prediction.gender == params.gender)
    if params.date_from is not None:
        query = query.filter(Prediction.created_at >= params.date_from)
    if params.date_to is not None:
        query = query.filter(Prediction.created_at <= params.date_to)
    if params.smoking_status is not None:
        query = query.filter(Prediction.smoking_status == params.smoking_status)
    if params.hypertension is not None:
        query = query.filter(Prediction.hypertension == params.hypertension)
    if params.heart_disease is not None:
        query = query.filter(Prediction.heart_disease == params.heart_disease)
    if params.residence_type is not None:
        query = query.filter(Prediction.Residence_type == params.residence_type)
    if params.work_type is not None:
        query = query.filter(Prediction.work_type == params.work_type)
    rows = query.order_by(Prediction.created_at.desc()).all()

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["PreStrokeNet export summary", "Value"])
    summary.append(["Generated at", datetime.now(timezone.utc).isoformat()])
    summary.append(["Rows exported", len(rows)])
    summary.append(["High risk rows", sum(1 for row in rows if (row.risk or "").lower() == "high")])
    summary.append(["Average probability", sum((row.final_probability or 0) for row in rows) / len(rows) if rows else 0])
    for cell in summary[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
    summary["B5"].number_format = "0.0%"
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 30

    predictions = workbook.create_sheet("Predictions")
    headers = _prediction_export_headers()
    predictions.append(headers)
    data = [_prediction_export_row(row) for row in rows]
    for row in data:
        predictions.append(row)
    _style_prediction_sheet(predictions, headers, data)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_csv(db: Session, params: PredictionSearchParams) -> bytes:
    result = search_predictions(db, params)
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(_prediction_export_headers())
    for item in result.items:
        writer.writerow([
            item.id,
            item.patient_name or "",
            item.patient_id or "",
            "",
            item.status or "draft",
            "",
            item.age if item.age is not None else "",
            item.gender if item.gender is not None else "",
            item.created_at.isoformat() if item.created_at else "",
            item.final_probability,
            item.clinical_probability,
            item.keystroke_probability,
            item.risk,
        ])
    return output.getvalue().encode("utf-8-sig")


def send_report_email(db: Session, prediction: Prediction, request: EmailReportRequest, actor_id: int | None = None) -> None:
    if not SMTP_HOST:
        raise RuntimeError("Report email is not configured. Configure SMTP_HOST before sending reports.")

    message = EmailMessage()
    message["Subject"] = request.subject
    message["From"] = SMTP_FROM_EMAIL or SMTP_USERNAME
    message["To"] = request.recipient
    message.set_content(
        f"{request.message}\n\nDiagnosis: {prediction.diagnosis or 'Not recorded'}\n"
        f"Recommendation: {prediction.recommendation or 'Not recorded'}\n"
        f"Follow-up date: {prediction.follow_up_date.isoformat() if prediction.follow_up_date else 'Not scheduled'}"
    )
    message.add_attachment(build_pdf(prediction), maintype="application", subtype="pdf", filename=f"PreStrokeNet_Prediction_{prediction.id}.pdf")

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as smtp:
            if SMTP_USE_TLS:
                smtp.starttls()
            if SMTP_USERNAME and SMTP_PASSWORD:
                smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
            smtp.send_message(message)
    except (OSError, smtplib.SMTPException) as error:
        raise RuntimeError("The report email could not be sent.") from error

    prediction.email_sent = True
    record_activity(db, activity_type="email_sent", message="Email sent", prediction_id=prediction.id, actor_id=actor_id)
    db.commit()
